import requests
from bs4 import BeautifulSoup
import pandas as pd
import nltk
from nltk.corpus import stopwords
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

nltk.download('stopwords')
stop_words = set(stopwords.words('english'))

def get_recipe_urls(base_url, max_pages=36):
    recipe_urls = []
    #in order to go through all the pages of the recipes
    for page in range(1, max_pages + 1):
        response = requests.get(f'{base_url}?fwp_paged={page}')
        if response.status_code == 200:
            soup = BeautifulSoup(response.content, 'html.parser')
            #get all the recipes links
            for link in soup.find_all('a', class_='entry-title-link'):
                recipe_urls.append(link['href'])
    return recipe_urls

def get_recipe_details(recipe_url):
    #get a recipe url and extract data about it.
    response = requests.get(recipe_url)
    if response.status_code == 200:
        soup = BeautifulSoup(response.content, 'html.parser')
        ingredients = []
        #get ingrediends data
        for ingredient in soup.find_all('li', class_='wprm-recipe-ingredient'):
            amount = ingredient.find('span', class_='wprm-recipe-ingredient-amount')
            name = ingredient.find('span', class_='wprm-recipe-ingredient-name')
            unit = ingredient.find('span', class_='wprm-recipe-ingredient-unit')

            amount_text = amount.get_text(strip=True) if amount else ''
            name_text = name.get_text(strip=True) if name else ''
            unit_text = unit.get_text(strip=True) if unit else ''

            #ingredient's full name, amount and unit
            full_ingredient = f"{amount_text} {unit_text} {name_text}".strip()
            full_ingredient = ' '.join(full_ingredient.split())  # Remove extra spaces
            ingredients.append(full_ingredient.lower())

        ingredients = '\n'.join(ingredients)

        recipe_name = soup.find('h1', class_='entry-title').get_text(strip=True) if soup.find('h1', class_='entry-title') else 'No title'

        #return a dictionary that contains all the data about the recipe
        return {
            'url': recipe_url,
            'name': recipe_name,
            'ingredients': ingredients
        }
    return None

def has_all_ingredients(ingredient_string, ingredients):
    for ingredient in ingredients:
        if ingredient.lower() not in ingredient_string:
            return False
    return True

def find_recipes_by_ingredients(base_url, ingredients, max_recipes=20, max_pages=36):
    recipe_urls = get_recipe_urls(base_url, max_pages)
    matching_recipes = []
    #go through all the URLS and if the recipe contains all the mentioned ingredients, save it's data
    for url in recipe_urls:
        #check if 20 recipes were found already
        if len(matching_recipes) >= max_recipes:
            break
        details = get_recipe_details(url)
        if details and has_all_ingredients(details['ingredients'], ingredients):
            matching_recipes.append(details)
    return matching_recipes

def save_recipes_to_excel(recipes, filename='recipes.xlsx'):
    df = pd.DataFrame(recipes, columns=['name', 'url', 'ingredients'])
    df.to_excel(filename, index=False)

    # Load the workbook and select the active worksheet
    wb = load_workbook(filename)
    ws = wb.active

    # In order to fit the width to the text
    header_font = Font(bold=True)
    for col in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col)
        ws[f'{col_letter}1'].font = header_font
        ws[f'{col_letter}1'].alignment = Alignment(horizontal='center', vertical='center')

    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws[f'{get_column_letter(col)}{row}'].alignment = Alignment(vertical='top', wrap_text=True)

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Save the workbook
    wb.save(filename)

def main():
    url = 'https://www.recipetineats.com/category/quick-and-easy/'
    ingreds = ['1 onion', '2 tbsp olive oil']
    recipes_for_query = find_recipes_by_ingredients(url, ingreds, max_recipes=20, max_pages=5)

    #if recipes with these ingrediens were found, save the data to the excel
    if recipes_for_query:
        save_recipes_to_excel(recipes_for_query)
        print(f'Found {len(recipes_for_query)} recipes. Check out recipes.xlsx')

    else:
        print('No recipes found matching the given ingredients.')

if __name__ == '__main__':
    main()
