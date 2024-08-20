import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

"""this file contains all three queries for the project:
 * all the recipes containing 1 onion and 2 tbsp olive oil
 * all the recipes published in 2024
 * all the recipes rated with 5 stars """


def get_recipe_urls(base_url, max_pages=36):
    recipe_urls = []
    for page in range(1, max_pages + 1):
        response = requests.get(f'{base_url}?fwp_paged={page}')
        if response.status_code == 200:
            soup = BeautifulSoup(response.content, 'html.parser')
            for link in soup.find_all('a', class_='entry-title-link'):
                recipe_urls.append(link['href'])
    return recipe_urls


def get_recipe_details(recipe_url):
    response = requests.get(recipe_url)
    if response.status_code == 200:
        soup = BeautifulSoup(response.content, 'html.parser')
        ingredients = []
        for ingredient in soup.find_all('li', class_='wprm-recipe-ingredient'):
            amount = ingredient.find('span', class_='wprm-recipe-ingredient-amount')
            name = ingredient.find('span', class_='wprm-recipe-ingredient-name')
            unit = ingredient.find('span', class_='wprm-recipe-ingredient-unit')

            amount_text = amount.get_text(strip=True) if amount else ''
            name_text = name.get_text(strip=True) if name else ''
            unit_text = unit.get_text(strip=True) if unit else ''

            full_ingredient = f"{amount_text} {unit_text} {name_text}".strip()
            full_ingredient = ' '.join(full_ingredient.split())  # Remove extra spaces
            ingredients.append(full_ingredient.lower())

        ingredients = '\n'.join(ingredients)

        recipe_name = soup.find('h1', class_='entry-title').get_text(strip=True) if soup.find('h1',
                                                                                              class_='entry-title') else 'No title'

        # Extract the date published
        date_published = soup.find('time', class_='entry-time').get_text(strip=True) if soup.find('time',
                                                                                                  class_='entry-time') else 'Unknown'

        # Extract rating
        rating = soup.find('span', class_='wprm-recipe-rating-average').get_text(strip=True) if soup.find('span',
                                                                                                          class_='wprm-recipe-rating-average') else 'No rating'

        return {
            'url': recipe_url,
            'name': recipe_name,
            'ingredients': ingredients,
            'date_published': date_published,
            'rating': rating
        }
    return None


def has_all_ingredients(ingredient_string, ingredients):
    for ingredient in ingredients:
        if ingredient.lower() not in ingredient_string:
            return False
    return True


def find_recipes_by_ingredients(base_url, ingredients, max_recipes=20, max_pages=36):
    recipe_urls = get_recipe_urls(base_url, max_pages)
    matching_recipes = []
    recipes_rated_5 = []
    recipes_from_24 = []
    for url in recipe_urls:
        if len(matching_recipes) >= max_recipes:
            break
        details = get_recipe_details(url)
        if details and has_all_ingredients(details['ingredients'], ingredients):
            matching_recipes.append(details)
        if details and details['rating'] == '5':
            recipes_rated_5.append({'url': details['url'], 'rating': details['rating']})
        if details and "24" in details['date_published']:
            recipes_from_24.append({'url': details['url'], 'date_published': details['date_published']})
    return matching_recipes, recipes_rated_5, recipes_from_24


def save_recipes_to_excel(recipes, filename, columns):
    df = pd.DataFrame(recipes, columns=columns)
    df.to_excel(filename, index=False)

    wb = load_workbook(filename)
    ws = wb.active

    header_font = Font(bold=True)
    for col in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col)
        ws[f'{col_letter}1'].font = header_font
        ws[f'{col_letter}1'].alignment = Alignment(horizontal='center', vertical='center')

    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws[f'{get_column_letter(col)}{row}'].alignment = Alignment(vertical='top', wrap_text=True)

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    wb.save(filename)


def main():
    url = 'https://www.recipetineats.com/category/quick-and-easy/'
    ingreds = ['1 onion', '2 tbsp olive oil']

    recipes_for_query, recipes_rated_5, recipes_from_24 = find_recipes_by_ingredients(url, ingreds, max_recipes=20,
                                                                                      max_pages=20)

    # Save recipes with ingredients to Excel (no rating or date_published)
    if recipes_for_query:
        save_recipes_to_excel(recipes_for_query, 'recipes_with_ingredients.xlsx',
                              columns=['name', 'url', 'ingredients'])
        print(
            f'Found {len(recipes_for_query)} recipes with specified ingredients. Check out recipes_with_ingredients.xlsx')
    else:
        print('No recipes found matching the given ingredients.')

    # Save recipes published in 2024 to Excel (only url and date_published)
    if recipes_from_24:
        save_recipes_to_excel(recipes_from_24, 'recipes_2024.xlsx', columns=['url', 'date_published'])
        print(f'Found {len(recipes_from_24)} recipes published in 2024. Check out top_10_recipes_2024.xlsx')
    else:
        print('No recipes found published in 2024.')

    # Save 5-star recipes to Excel (only url and rating)
    if recipes_rated_5:
        save_recipes_to_excel(recipes_rated_5, 'five_star_recipes.xlsx', columns=['url', 'rating'])
        print(f'Found {len(recipes_rated_5)} 5-star recipes. Check out top_10_five_star_recipes.xlsx')
    else:
        print('No 5-star recipes found.')


if __name__ == '__main__':
    main()
